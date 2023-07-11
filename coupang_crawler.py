import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import sys

"""
    Usage
        - 파이썬 파일이 위치한 경로에 data 폴더를 만들고, coupang_url.txt 파일을 생성합니다.
        - coupang_url.txt 파일의 각 줄에는 담고자하는 상품의 url과 수량을 띄어쓰기로 구분하여 작성합니다. ex) https://coupang.com/robotvision 3
        - python coupang_crawler.py 로 실행하면 data 폴더에 coupang.xlsx 엑셀 파일이 생성됩니다.
"""

class coupang_crawler():
    def __init__(self):
        self.headers = {"authority": "www.coupang.com",
            "method": "GET",
            "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
            "accept-encoding": "gzip, deflate, br",
            "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.104 Whale/3.13.131.36 Safari/537.36",
            "sec-ch-ua-platform": "macOS",
            "cookie": "PCID=31489593180081104183684; _fbp=fb.1.1644931520418.1544640325; gd1=Y; X-CP-PT-locale=ko_KR; MARKETID=31489593180081104183684; sid=03ae1c0ed61946c19e760cf1a3d9317d808aca8b; x-coupang-origin-region=KOREA; x-coupang-target-market=KR; x-coupang-accept-language=ko_KR;"}

        # # 크롤링할 웹 페이지 URL txt 파일 읽기
        # file = open("./data/coupang_url.txt", 'r')

        self.x = openpyxl.Workbook()
        self.WS = self.x.active

        self.align_center = Alignment(horizontal='center', vertical='center')
        self.align_right = Alignment(horizontal='right', vertical='center')
        self.subtitle_color = PatternFill('solid', fgColor='C8C8FF')

        self.WS['B2'].value = '제품명'
        self.WS['B2'].alignment = self.align_center
        self.WS['B2'].fill = self.subtitle_color
        self.WS['C2'].value = '수량'
        self.WS['C2'].alignment = self.align_center
        self.WS['C2'].fill = self.subtitle_color
        self.WS['D2'].value = '링크'
        self.WS['D2'].alignment = self.align_center
        self.WS['D2'].fill = self.subtitle_color
        self.WS['E2'].value = '가격'
        self.WS['E2'].alignment = self.align_center
        self.WS['E2'].fill = self.subtitle_color

    def get_instance(self, url, quantity):
        # requests 모듈을 사용하여 웹 페이지의 HTML 코드 가져오기
        html = requests.get(url=url, headers=self.headers).content

        # BeautifulSoup을 사용하여 HTML 코드 파싱
        soup = BeautifulSoup(html, 'html.parser')

        product_name = soup.find(attrs={"class", "prod-buy-header__title"}).get_text()
        prices = []
        for strong in soup.find_all("strong"):
            for child in strong.children:
                try:
                    if child[-1] == "원" and len(child) > 1:
                        prices.append(int(child[:-1].replace(',','')))
                except:
                    continue
        price = str(min(prices))

        return product_name, quantity, price
    
    def save_pyxl(self, instances, total_price):
        for index, instance in enumerate(instances):
            index += 3

            self.WS['B{}'.format(index)].value = instance["product_name"]
            self.WS['C{}'.format(index)].value = instance["quantity"]
            self.WS['C{}'.format(index)].alignment = self.align_center
            self.WS['D{}'.format(index)].value = instance["link"]
            self.WS['E{}'.format(index)].value = "{0:,d}원".format(instance["price"])
            self.WS['E{}'.format(index)].alignment = self.align_center

        self.WS.column_dimensions['B'].width = 70
        self.WS.column_dimensions['C'].width = 10
        self.WS.column_dimensions['D'].width = 100
        self.WS.column_dimensions['E'].width = 20
        self.WS['D{}'.format(len(instances)+3)].value = "총계   "
        self.WS['D{}'.format(len(instances)+3)].alignment = self.align_right
        self.WS['E{}'.format(len(instances)+3)].value = format(total_price, ',') + "원"
        self.WS['E{}'.format(len(instances)+3)].alignment = self.align_center
        self.x.save('./data/coupang.xlsx')

        # self.WS['B{}'.format(index)].value = product_name
        # self.WS['C{}'.format(index)].value = ea
        # self.WS['C{}'.format(index)].alignment = align_center
        # self.WS['D{}'.format(index)].value = url
        # self.WS['E{}'.format(index)].value = price
        # self.WS['E{}'.format(index)].alignment = align_center

# while True:
#     # 크롤링할 웹 페이지 URL
#     data = file.readline().split()
#     if not data: break
#     url = data[0]
#     ea = data[1]

#     # requests 모듈을 사용하여 웹 페이지의 HTML 코드 가져오기
#     html = requests.get(url=url, headers=headers).content

#     # BeautifulSoup을 사용하여 HTML 코드 파싱
#     soup = BeautifulSoup(html, 'html.parser')

#     product_name = soup.find(attrs={"class", "prod-buy-header__title"}).get_text()
#     price = soup.strong.get_text()

#     self.WS['B{}'.format(index)].value = product_name
#     self.WS['C{}'.format(index)].value = ea
#     self.WS['C{}'.format(index)].alignment = align_center
#     self.WS['D{}'.format(index)].value = url
#     self.WS['E{}'.format(index)].value = price
#     self.WS['E{}'.format(index)].alignment = align_center

#     index += 1
#     total_price += int(price[:-1].replace(',','')) * int(ea)

# self.WS.column_dimensions['B'].width = 70
# self.WS.column_dimensions['C'].width = 10
# self.WS.column_dimensions['D'].width = 100
# self.WS.column_dimensions['E'].width = 20
# self.WS['D{}'.format(index+1)].value = "총계   "
# self.WS['D{}'.format(index+1)].alignment = align_right
# self.WS['E{}'.format(index+1)].value = format(total_price, ',') + "원"
# self.WS['E{}'.format(index+1)].alignment = align_center
# self.x.save('./data/coupang.xlsx')